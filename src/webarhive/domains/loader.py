"""Domain loading from various sources (spec §2.3).

Sources: manual single, clipboard paste (many lines), Excel, CSV, TXT.
For Excel/CSV the domains come from the FIRST column (fixed format).
For TXT — one per line.

Output always passes through normalize_domain and dedup. A LoadReport
is returned for the "loaded X → valid unique Y → dropped Z" summary
(spec §2.4 footer).
"""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field

from webarhive.domains.normalize import NormalizeResult, normalize_domain


@dataclass
class LoadReport:
    raw_lines: int = 0
    valid_unique: list[str] = field(default_factory=list)
    rejected: list[NormalizeResult] = field(default_factory=list)

    @property
    def dropped(self) -> int:
        return self.raw_lines - len(self.valid_unique)


def _normalize_many(raw_items: list[str], *, check_subdomains: bool) -> LoadReport:
    report = LoadReport()
    seen: set[str] = set()
    for raw in raw_items:
        if raw is None:
            continue
        s = str(raw).strip()
        if not s:
            continue
        report.raw_lines += 1
        res = normalize_domain(s, check_subdomains=check_subdomains)
        if not res.ok or not res.domain:
            report.rejected.append(res)
            continue
        if res.domain in seen:
            continue
        seen.add(res.domain)
        report.valid_unique.append(res.domain)
    return report


def load_from_text(text: str, *, check_subdomains: bool = False) -> LoadReport:
    """Parse a free-form blob (clipboard paste, manual textarea).

    Splits on newlines, commas, semicolons, tabs and whitespace runs.
    """
    if not text:
        return LoadReport()
    # Normalize separators to newline, then split.
    for sep in (",", ";", "\t"):
        text = text.replace(sep, "\n")
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    return _normalize_many(lines, check_subdomains=check_subdomains)


def _load_csv(data: bytes, *, check_subdomains: bool) -> LoadReport:
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    items: list[str] = []
    for row in reader:
        if not row:
            continue
        items.append(row[0])
    return _normalize_many(items, check_subdomains=check_subdomains)


def _load_xlsx(data: bytes, *, check_subdomains: bool) -> LoadReport:
    from openpyxl import load_workbook  # heavy import, lazy

    wb = load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    items: list[str] = []
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        cell = row[0]
        if cell is None:
            continue
        items.append(str(cell))
    wb.close()
    return _normalize_many(items, check_subdomains=check_subdomains)


def _load_txt(data: bytes, *, check_subdomains: bool) -> LoadReport:
    text = data.decode("utf-8", errors="replace")
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    return _normalize_many(lines, check_subdomains=check_subdomains)


def load_from_bytes(filename: str, data: bytes, *, check_subdomains: bool = False) -> LoadReport:
    """Detect format from filename suffix and parse.

    Supported: .xlsx, .csv, .txt. Anything else is treated as TXT.
    """
    name = filename.lower()
    if name.endswith(".xlsx"):
        return _load_xlsx(data, check_subdomains=check_subdomains)
    if name.endswith(".csv"):
        return _load_csv(data, check_subdomains=check_subdomains)
    return _load_txt(data, check_subdomains=check_subdomains)
